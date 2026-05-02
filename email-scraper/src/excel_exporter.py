"""
Excel Exporter module for exporting extracted data to Excel and CSV formats.
"""

import logging
from pathlib import Path
from typing import List, Dict, Optional
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Export extracted order data to Excel and CSV formats."""
    
    def __init__(self, config):
        self.config = config
        
    def export_to_excel(self, orders_data: List[Dict], output_file: str = None) -> Optional[str]:
        """
        Export order data to an Excel file with multiple sheets.
        
        Args:
            orders_data: List of order dictionaries
            output_file: Output filename (default from config)
            
        Returns:
            Path to created file or None if failed
        """
        if not orders_data:
            logger.warning("No data to export")
            return None
        
        if output_file is None:
            output_file = self.config.output_excel_file
        
        output_path = Path(output_file)
        
        try:
            logger.info(f"Creating Excel file: {output_path}")
            
            # Create workbook
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            
            # Create summary sheet
            self._create_order_summary_sheet(wb, orders_data)
            
            # Create line items sheet
            self._create_line_items_sheet(wb, orders_data)
            
            # Create processing summary sheet
            self._create_processing_summary_sheet(wb, orders_data)
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"Excel file saved: {output_path} ({len(orders_data)} orders)")
            
            # Also export to CSV
            self.export_to_csv(orders_data)
            
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Error creating Excel file: {str(e)}")
            return None
    
    def _create_order_summary_sheet(self, wb: Workbook, orders_data: List[Dict]):
        """Create the order summary sheet."""
        ws = wb.create_sheet("Order Summary")
        
        # Headers
        headers = [
            "Order ID", "Vendor", "Invoice Number", "Invoice Date",
            "Total Amount", "Currency", "Email From", "Email Subject",
            "Processing Date", "Source File", "Line Items Count"
        ]
        
        # Style for header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data rows
        for row_idx, order in enumerate(orders_data, 2):
            ws.cell(row=row_idx, column=1, value=order.get('id', ''))
            ws.cell(row=row_idx, column=2, value=order.get('vendor', ''))
            ws.cell(row=row_idx, column=3, value=order.get('invoice_number', ''))
            ws.cell(row=row_idx, column=4, value=order.get('invoice_date', ''))
            ws.cell(row=row_idx, column=5, value=order.get('total_amount'))
            ws.cell(row=row_idx, column=6, value=order.get('currency', 'USD'))
            ws.cell(row=row_idx, column=7, value=order.get('email_from', ''))
            ws.cell(row=row_idx, column=8, value=order.get('email_subject', ''))
            ws.cell(row=row_idx, column=9, value=order.get('extraction_date', '')[:10] if order.get('extraction_date') else '')
            ws.cell(row=row_idx, column=10, value=order.get('source_file', ''))
            ws.cell(row=row_idx, column=11, value=len(order.get('line_items', [])))
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws, headers)
        
        # Add filters
        ws.auto_filter.ref = ws.dimensions
    
    def _create_line_items_sheet(self, wb: Workbook, orders_data: List[Dict]):
        """Create the line items detail sheet."""
        ws = wb.create_sheet("Line Items")
        
        # Headers
        headers = [
            "Order ID", "Vendor", "Invoice Number", "Item #",
            "Description", "Quantity", "Unit Price", "Amount"
        ]
        
        # Style for header row
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data rows
        row_idx = 2
        for order in orders_data:
            line_items = order.get('line_items', [])
            for item_idx, item in enumerate(line_items, 1):
                ws.cell(row=row_idx, column=1, value=order.get('id', ''))
                ws.cell(row=row_idx, column=2, value=order.get('vendor', ''))
                ws.cell(row=row_idx, column=3, value=order.get('invoice_number', ''))
                ws.cell(row=row_idx, column=4, value=item_idx)
                ws.cell(row=row_idx, column=5, value=item.get('description', ''))
                ws.cell(row=row_idx, column=6, value=item.get('quantity'))
                ws.cell(row=row_idx, column=7, value=item.get('unit_price'))
                ws.cell(row=row_idx, column=8, value=item.get('amount'))
                row_idx += 1
        
        # Auto-adjust column widths
        self._adjust_column_widths(ws, headers)
        
        # Add filters
        ws.auto_filter.ref = ws.dimensions
    
    def _create_processing_summary_sheet(self, wb: Workbook, orders_data: List[Dict]):
        """Create a processing summary sheet with statistics."""
        ws = wb.create_sheet("Processing Summary")
        
        # Calculate statistics
        total_orders = len(orders_data)
        total_line_items = sum(len(order.get('line_items', [])) for order in orders_data)
        total_amount = sum(order.get('total_amount', 0) or 0 for order in orders_data)
        
        orders_with_invoices = sum(1 for order in orders_data if order.get('invoice_number'))
        orders_with_totals = sum(1 for order in orders_data if order.get('total_amount'))
        
        # Vendors summary
        vendor_counts = {}
        for order in orders_data:
            vendor = order.get('vendor', 'Unknown')
            vendor_counts[vendor] = vendor_counts.get(vendor, 0) + 1
        
        # Write summary
        ws.merge_cells('A1:B1')
        ws.cell(row=1, column=1, value="Processing Summary Report").font = Font(bold=True, size=14)
        
        summary_data = [
            ("Total Orders Processed:", total_orders),
            ("Total Line Items:", total_line_items),
            ("Total Amount:", f"${total_amount:,.2f}"),
            ("Orders with Invoice Numbers:", orders_with_invoices),
            ("Orders with Total Amounts:", orders_with_totals),
            ("Processing Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ]
        
        for row_idx, (label, value) in enumerate(summary_data, 3):
            ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row_idx, column=2, value=value)
        
        # Vendor breakdown
        if vendor_counts:
            ws.cell(row=12, column=1, value="Vendor Breakdown").font = Font(bold=True, size=12)
            ws.cell(row=13, column=1, value="Vendor").font = Font(bold=True)
            ws.cell(row=13, column=2, value="Order Count").font = Font(bold=True)
            
            for row_idx, (vendor, count) in enumerate(sorted(vendor_counts.items(), key=lambda x: x[1], reverse=True), 14):
                ws.cell(row=row_idx, column=1, value=vendor)
                ws.cell(row=row_idx, column=2, value=count)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
    
    def _adjust_column_widths(self, ws, headers: List[str]):
        """Auto-adjust column widths based on content."""
        for col_idx, header in enumerate(headers, 1):
            max_length = len(str(header))
            
            for row in ws.iter_rows(min_row=2, max_col=col_idx, max_row=ws.max_row):
                cell_value = row[col_idx - 1].value
                if cell_value:
                    max_length = max(max_length, min(len(str(cell_value)), 50))
            
            ws.column_dimensions[chr(64 + col_idx)].width = max_length + 2
    
    def export_to_csv(self, orders_data: List[Dict]) -> Optional[str]:
        """
        Export order data to CSV files.
        
        Args:
            orders_data: List of order dictionaries
            
        Returns:
            Path to CSV directory or None if failed
        """
        if not orders_data:
            return None
        
        csv_dir = self.config.csv_export_dir
        csv_dir.mkdir(parents=True, exist_ok=True)
        
        try:
            # Export order summary CSV
            summary_data = []
            for order in orders_data:
                summary_data.append({
                    'order_id': order.get('id', ''),
                    'vendor': order.get('vendor', ''),
                    'invoice_number': order.get('invoice_number', ''),
                    'invoice_date': order.get('invoice_date', ''),
                    'total_amount': order.get('total_amount'),
                    'currency': order.get('currency', 'USD'),
                    'email_from': order.get('email_from', ''),
                    'email_subject': order.get('email_subject', ''),
                    'processing_date': order.get('extraction_date', ''),
                    'source_file': order.get('source_file', ''),
                    'line_items_count': len(order.get('line_items', []))
                })
            
            summary_df = pd.DataFrame(summary_data)
            summary_csv = csv_dir / 'orders_summary.csv'
            summary_df.to_csv(summary_csv, index=False)
            logger.info(f"Created CSV: {summary_csv}")
            
            # Export line items CSV
            line_items_data = []
            for order in orders_data:
                for item_idx, item in enumerate(order.get('line_items', []), 1):
                    line_items_data.append({
                        'order_id': order.get('id', ''),
                        'vendor': order.get('vendor', ''),
                        'invoice_number': order.get('invoice_number', ''),
                        'item_number': item_idx,
                        'description': item.get('description', ''),
                        'quantity': item.get('quantity'),
                        'unit_price': item.get('unit_price'),
                        'amount': item.get('amount')
                    })
            
            if line_items_data:
                line_items_df = pd.DataFrame(line_items_data)
                line_items_csv = csv_dir / 'line_items.csv'
                line_items_df.to_csv(line_items_csv, index=False)
                logger.info(f"Created CSV: {line_items_csv}")
            
            return str(csv_dir)
            
        except Exception as e:
            logger.error(f"Error exporting to CSV: {str(e)}")
            return None
    
    def create_demo_data(self) -> List[Dict]:
        """
        Create demo data for testing without email connection.
        
        Returns:
            List of demo order dictionaries
        """
        demo_orders = [
            {
                'id': 'DEMO-001',
                'vendor': 'ABC Supplies Inc.',
                'invoice_number': 'INV-2024-001',
                'invoice_date': '2024-01-15',
                'total_amount': 1250.50,
                'currency': 'USD',
                'email_from': 'sales@abcsupplies.com',
                'email_subject': 'Order Confirmation - INV-2024-001',
                'extraction_date': datetime.now().isoformat(),
                'source_file': 'invoice_001.pdf',
                'line_items': [
                    {
                        'description': 'Office Paper A4 (Box)',
                        'quantity': 10,
                        'unit_price': 25.50,
                        'amount': 255.00
                    },
                    {
                        'description': 'Ballpoint Pens (Pack of 50)',
                        'quantity': 5,
                        'unit_price': 15.00,
                        'amount': 75.00
                    },
                    {
                        'description': 'Desk Organizer Set',
                        'quantity': 8,
                        'unit_price': 45.00,
                        'amount': 360.00
                    }
                ]
            },
            {
                'id': 'DEMO-002',
                'vendor': 'TechGear Solutions',
                'invoice_number': 'TG-2024-0156',
                'invoice_date': '2024-01-18',
                'total_amount': 3499.99,
                'currency': 'USD',
                'email_from': 'orders@techgear.com',
                'email_subject': 'Your Order #TG-2024-0156',
                'extraction_date': datetime.now().isoformat(),
                'source_file': 'invoice_002.pdf',
                'line_items': [
                    {
                        'description': 'Wireless Mouse',
                        'quantity': 25,
                        'unit_price': 29.99,
                        'amount': 749.75
                    },
                    {
                        'description': 'USB-C Hub Adapter',
                        'quantity': 15,
                        'unit_price': 49.99,
                        'amount': 749.85
                    },
                    {
                        'description': 'Mechanical Keyboard',
                        'quantity': 10,
                        'unit_price': 89.99,
                        'amount': 899.90
                    },
                    {
                        'description': 'Monitor Stand Adjustable',
                        'quantity': 12,
                        'unit_price': 79.99,
                        'amount': 959.88
                    }
                ]
            }
        ]
        
        logger.info(f"Created {len(demo_orders)} demo orders")
        return demo_orders
